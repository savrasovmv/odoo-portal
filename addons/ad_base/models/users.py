# -*- coding: utf-8 -*-

from odoo import fields, models, api
from ldap3 import Server, Connection, SUBTREE, MODIFY_REPLACE, LEVEL
import json
import base64
import re

from openpyxl import Workbook
import os, fnmatch
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

class AdOrganization(models.Model):
    _name = "ad.organizacion"
    _description = "Организации AD"
    _order = "name"

    name = fields.Char(u'Наименование', required=True)
    active = fields.Boolean('Active', default=True)
   

class AdBranch(models.Model):
    _name = "ad.branch"
    _description = "Подразделение AD"
    _order = "name"

    name = fields.Char(u'Наименование 1С', required=True)
    ad_name = fields.Char(u'Наименование в AD')
    adbook_name = fields.Char(u'Наименование в справочнике')
    # organization_id = fields.Many2one("ad.organizacion", string="Организация")
    company_id = fields.Many2one('res.company', string='Компания')

    active = fields.Boolean('Active', default=True)
    address = fields.Text(string='Адрес')
    is_view_adbook = fields.Boolean(string='Отоброжать в справочнике контактов')
    is_view_photo_adbook = fields.Boolean(string='Показывать фото в справочнике контактов')
    sequence = fields.Integer(string=u"Сортировка", help="Сортировка", default=10)

    is_default_branch = fields.Boolean(string='AD контейнер по умолчанию', help='Если установлено, новые пользователи не вошедшие не в одну группу будут создаваться тут')
    hr_department_id = fields.Many2one("hr.department", string="HR Подразделене")



class AdDepartment(models.Model):
    _name = "ad.department"
    _description = "Управления/отделы AD"
    _order = "sequence"

    name = fields.Char(u'Наименование', required=True)
    branch_id = fields.Many2one("ad.branch", string="Подразделение")
    active = fields.Boolean('Active', default=True)
    is_view_adbook = fields.Boolean(string='Отоброжать в справочнике контактов', default=True)
    sequence = fields.Integer(string=u"Сортировка", help="Сортировка", default=10)

class AdGroup(models.Model):
    _name = "ad.group"
    _description = "Группы AD"
    _order = "name"

    name = fields.Char(u'Наименование', required=True)

    distinguished_name = fields.Char(u'AD distinguishedName')
    account_name = fields.Char(u'sAMAccountName')
    object_SID = fields.Char(u'AD objectSID')
    is_ldap = fields.Boolean('LDAP?', default=False)

    active = fields.Boolean('Active', default=True)
    is_managed = fields.Boolean('Управляемая', default=False, help="Включите, для управления вхождения пользователя в эту группу в форме Пользователя")



flags = [
    [0x0001, 'SCRIPT'],
    [0x0002, 'ACCOUNTDISABLE'],
    [0x0008, 'HOMEDIR_REQUIRED'],
    [0x0010, 'LOCKOUT'],
    [0x0020, 'PASSWD_NOTREQD'],
    [0x0040, 'PASSWD_CANT_CHANGE'],
    [0x0080, 'ENCRYPTED_TEXT_PWD_ALLOWED'],
    [0x0100, 'TEMP_DUPLICATE_ACCOUNT'],
    [0x0200, 'NORMAL_ACCOUNT'],
    [0x0800, 'INTERDOMAIN_TRUST_ACCOUNT'],
    [0x1000, 'WORKSTATION_TRUST_ACCOUNT'],
    [0x2000, 'SERVER_TRUST_ACCOUNT'],
    [0x10000, 'DONT_EXPIRE_PASSWORD'],
    [0x20000, 'MNS_LOGON_ACCOUNT'],
    [0x40000, 'SMARTCARD_REQUIRED'],
    [0x80000, 'TRUSTED_FOR_DELEGATION'],
    [0x100000, 'NOT_DELEGATED'],
    [0x200000, 'USE_DES_KEY_ONLY'],
    [0x400000, 'DONT_REQ_PREAUTH'],
    [0x800000, 'PASSWORD_EXPIRED'],
    [0x1000000, 'TRUSTED_TO_AUTH_FOR_DELEGATION'],
    [0x04000000, 'PARTIAL_SECRETS_ACCOUNT'],
  ]


class AdUsers(models.Model):
    _name = "ad.users"
    _description = "Пользователи AD"
    _order = "name"

    name = fields.Char(u'ФИО', required=True)
    employee_id = fields.Many2one("hr.employee", string="Сотрудник")

    active = fields.Boolean('Active', default=True)
    is_ldap = fields.Boolean('LDAP?', default=False)

    # organization_id = fields.Many2one("ad.organizacion", string="Организация", compute="_compute_organization", store=True)
    company_id = fields.Many2one('res.company', string='Компания', compute="_compute_company", store=True)

    branch_id = fields.Many2one("ad.branch", string="Подразделение")
    department_id = fields.Many2one("ad.department", string="Управление/отдел")
    title = fields.Char(u'Должность')

    ip_phone = fields.Char(u'Вн. номер')
    phone = fields.Char(u'Мобильный телефон 1')
    sec_phone = fields.Char(u'Мобильный телефон 2')

    email = fields.Char(u'E-mail')

    search_text = fields.Char(u'Поисковое поле', compute="_get_search_text", store=True, index=True)

    username = fields.Char(u'sAMAccountName')
    object_SID = fields.Char(u'AD objectSID')
    distinguished_name = fields.Char(u'AD distinguishedName')
    user_account_control = fields.Char(u'AD userAccountControl')
    user_account_control_result = fields.Char(u'AD userAccountControl result', compute="_get_user_account_control_result")

    photo = fields.Binary('Фото', default=None)

    birthday = fields.Date(string='Дата рождения')


    is_view_adbook = fields.Boolean(string='Отоброжать в справочнике контактов', default=True)
    is_view_disabled_adbook = fields.Boolean(string='Отоброжать в справочнике контактов даже если отключена', default=False)
    sequence = fields.Integer(string=u"Сортировка", help="Сортировка", default=10)


    # is_fired = fields.Boolean(string='Уволен', default=False)
    # fired_date = fields.Date(string='Дата увольнения')

    # is_vacation = fields.Boolean(string='Отпуск')
    # vacation_start_date = fields.Date(string='Дата начала отпуска')
    # vacation_end_date = fields.Date(string='Дата окончания отпуска')

    # is_btrip = fields.Boolean(string='Командировка')
    # btrip_start_date = fields.Date(string='Дата начала командировки')
    # btrip_end_date = fields.Date(string='Дата окончания командировки')

    #Блокировки
    is_yaware = fields.Boolean(string='yaware')
    is_usb_block = fields.Boolean(string='Блокировка USB')
    is_mailarchiva = fields.Boolean(string='mailarchiva')
    is_phone_rec = fields.Boolean(string='Запись телефонных разговоров')
    is_socnet_block = fields.Boolean(string='Блокировка соц.сетей')
    is_mesg_block = fields.Boolean(string='Блокировка мессенджеров')
    is_cloud_block = fields.Boolean(string='Блокировка облаков')
    is_email_block = fields.Boolean(string='Блокировка email')
    is_rem_ad_block = fields.Boolean(string='rem_ad_block')
    is_iw = fields.Boolean(string='iw')
    is_backup = fields.Boolean(string='backup')
    is_vpn = fields.Boolean(string='vpn')
    is_vip = fields.Boolean(string='vip')

    users_group_line = fields.One2many('ad.users_group_line', 'users_id', string=u"Строка Группы AD")


    # @api.depends("is_yaware")
    # def _is_yaware_yaware(self):
    #     self.is_usb_block = self.is_yaware
    @api.depends("name", "ip_phone", "phone","sec_phone", "email")
    def _get_search_text(self):
        for user in self:
            ip_phone = re.sub('\D', '', user.ip_phone) if user.ip_phone else ''
            phone = re.sub('\D', '', user.phone) if user.phone else ''
            sec_phone = re.sub('\D', '', user.sec_phone) if user.sec_phone else ''
            email = user.email if user.email else ''
            user.search_text = user.name + " " + ip_phone + " " + phone  + " " + sec_phone + " " + email


    def get_employee_by_name(self):
        for user in self:
            empl = self.env['hr.employee'].search([
                ('name', '=', user.name),
                '|',
                ('active', '=', False), 
                ('active', '=', True)

            ], limit=1)

            if len(empl)>0:
                user.employee_id = empl[0].id
                user.birthday = empl[0].birthday
                employee = self.env['hr.employee'].browse(empl[0].id)
                if user.photo:
                    employee.image_1920 = user.photo
                # else:
                #     employee.image_1920 = employee._default_image()
                employee.users_id = user.id
                employee.mobile_phone = user.phone
                employee.mobile_phone2 = user.sec_phone
                employee.ip_phone = user.ip_phone
                employee.work_email = user.email

            else:
                user.employee_id = None

    
    def join_user_and_employee(self, full_sync=False):
        for user in self.search([]):
            if not user.employee_id or full_sync:
                user.get_employee_by_name()



    def update_group_list(self):
        for empl in self:
            group_list = self.env['ad.group'].search([
                                                ('active', '=', True),
                                                ('is_managed', '=', True),
                                                ], order="name")
            
            empl.users_group_line.unlink()

            for group in group_list:

                empl.users_group_line.create({
                    'name': group.name,
                    'group_id': group.id,
                    'users_id': empl.id,
                })





    @api.depends("branch_id", "branch_id.organization_id")
    def _compute_organization(self):
        for record in self:
            if record.branch_id:
                record.organization_id = record.branch_id.organization_id
    
    @api.depends("branch_id", "branch_id.company_id")
    def _compute_company(self):
        for record in self:
            if record.branch_id:
                record.company_id = record.branch_id.company_id

    def _get_user_account_control_result(self):
        for record in self:
            if record.user_account_control:
                result = ''
                for value in flags:
                    if (int(record.user_account_control) | int(value[0])) == int(record.user_account_control):
                        result += value[1] + ','
                record.user_account_control_result = result

    def action_update_from_ldap(self):
        pass


    


    #Экспорт справочника в Excel
    def _export_adbook(self):
        print("+++ _export_adbook")

        #Стили
        bd = Side(style='thick', color="000000")#Жирные границы ячейки
        bb = Side(style='thin', color="000000")#Стандартные границы ячеек

        #Стиль для заголовков
        def top_cell_style(cell, top=False):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText = True)
            cell.border = Border(left=bb, top=bb, right=bb, bottom=bb)
            cell.fill = PatternFill('solid', fgColor='A8353A') if top else PatternFill('solid', fgColor='92CEFC')
            cell.font = Font(b=True, color='FFFFFF') if top else Font(b=True, color='4D4D4D')

        #Стиль для обычных ячеек
        def basic_cell_style(cell, ip_phone=False):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText = True) if ip_phone else Alignment(horizontal="left", vertical="center", wrapText = True)
            cell.border = Border(left=bb, top=bb, right=bb, bottom=bb)
            cell.fill = PatternFill('solid', fgColor='FFFFFF')
            

        file_name = '/tmp/ETS_Contacts.xlsx'

        wb = Workbook()

        branch_list = self.env['ad.branch'].search([
                                ('active', '=', True),
                                ('is_view_adbook', '=', True),
                            ], order="sequence desc")
        
        for branch in branch_list:

            #Создания листа
            ws = wb.create_sheet(branch.adbook_name  or branch.name )

            # Шапка 
            ws['A1'] = 'ФИО'
            ws.column_dimensions['A'].width = 45
            ws['B1'] = 'Должность'
            ws.column_dimensions['B'].width = 65.10
            ws['C1'] = 'Внутренний номер'
            ws.column_dimensions['C'].width = 24.43
            ws['D1'] = 'Мобильный телефон 1'
            ws.column_dimensions['D'].width = 24.43
            ws['E1'] = 'Мобильный телефон 2'
            ws.column_dimensions['E'].width = 24.43
            ws['F1'] = 'Электронная почта'
            ws.column_dimensions['F'].width = 31.29
            ws.merge_cells('A2:F2')
            for c in 'ABCDEF':
                top_cell_style(ws['%s1' % c], True)
                top_cell_style(ws['%s2' % c])
            ws.row_dimensions[2].height = 75
            ws['A2'] = branch.address or ''
            
            
            #Получаем список управлений/отделов 
            department_list_id = self.env['ad.users'].read_group([ 
                                                        ('branch_id', '=', branch.id),
                                                        ('active', '=', True),
                                                        ('is_view_adbook', '=', True),
                                                    ], 
                                                        fields=['department_id'], 
                                                        groupby=['department_id']
                                                    )
            #Список id отделв для поиска Пользователей ADов
            department_ids = []
            
            for data in department_list_id:
                d_id, obj = data['department_id']
                department_ids.append(d_id)

            department_list = self.env['ad.department'].search([ 
                                                    ('id', 'in', department_ids),
                                                    ('active', '=', True),
                                                    ('is_view_adbook', '=', True),
                                                ], 
                                                    order="sequence desc"
                                                )
            
            i = 3
            for department in department_list:
                
                ws.merge_cells('A%d:F%d' % (i, i))
                
                for c in 'ABCDEF':
                    top_cell_style(ws['%s%d' % (c, i)])
                    
                ws['A%d' % i] = department.name or ''
                
                i += 1
                
                users_list = self.env['ad.users'].search([
                            ('branch_id', '=', branch.id),
                            ('department_id', '=', department.id),
                            ('active', '=', True),
                            ('is_view_adbook', '=', True),
                        ], order="sequence desc")

                for user in users_list:
                    
                    ws['A%d' % i] = user.name or ''
                    basic_cell_style(ws['A%d' % i])
                    ws['B%d' % i] = user.title or ''
                    basic_cell_style(ws['B%d' % i])
                    ws['C%d' % i] = user.ip_phone or ''
                    basic_cell_style(ws['C%d' % i], True)
                    ws['D%d' % i] = user.phone or ''
                    basic_cell_style(ws['D%d' % i])
                    ws['E%d' % i] = user.sec_phone or ''
                    basic_cell_style(ws['E%d' % i])
                    ws['F%d' % i] = user.email or ''
                    basic_cell_style(ws['F%d' % i])

                    # if len(str(users.title)) > 60:
                    #     ws.row_dimensions[i].height = 35
                    

                    i += 1
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToHeight = False
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            #ws.print_rows = "1:1"
            #ws.print_title_cols = 'A:F'
            #ws.print_title_rows = '1:2'
            #ws.print_title_cols = 'A:F'
            #ws.print_options.horizontalCentered = True
            ws.print_area = 'A1:F%s' % i
                    
        
        
        #Удаляем страницу по умолчанию 
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        
        wb.save(filename = file_name)
        wb.close()

        return file_name

    
class UsersGroupLine(models.Model):
    _name = "ad.users_group_line"
    _description = "Строка Установка групп пользователя"
    _order = "name"

    name = fields.Char(u'Наименование', compute="_get_name")
    group_id = fields.Many2one("ad.group", string="Группа AD")
    is_enable = fields.Boolean(string='Включена?')

    users_id = fields.Many2one('ad.users',
		ondelete='cascade', string=u"Пользователь", required=True)

    @api.depends("group_id")
    def _get_name(self):
        self.name = self.group_id.name

    




